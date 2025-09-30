from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from difflib import SequenceMatcher
import difflib
import re


class FileComparer:
    def __init__(self, file1, file2, output_filename):
        self.file1 = file1
        self.file2 = file2
        self.output_filename = output_filename
        self.highlight_color_excel = 'FFFF00'
        self.text_color_excel = 'FF0000'
        self.highlight_color_html = '#FFD700'
        self.text_color_html = '#FF0000'

    @staticmethod
    def read_file(filename):
        with open(filename, 'r') as file:
            return file.readlines()

    def generate_excel_diff(self, file1_lines, file2_lines):
        # Create Workbook and select active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "File Comparison"

        # Define fill color for differences
        fill_diff = PatternFill(start_color=self.highlight_color_excel, end_color=self.highlight_color_excel, fill_type="solid")

        # Define font color for differences
        font_diff = Font(color=self.text_color_excel)

        # Initialize variables
        row = 1
        # Write headers
        ws.cell(row=row, column=1).value = 'Line'
        ws.cell(row=row, column=2).value = 'Device 1'
        ws.cell(row=row, column=3).value = 'Line'
        ws.cell(row=row, column=4).value = 'Device 2'
        row += 1

        # Use SequenceMatcher to find matching blocks
        matcher = SequenceMatcher(None, file1_lines, file2_lines)
        opcodes = matcher.get_opcodes()

        lineno1 = lineno2 = 0

        for tag, i1, i2, j1, j2 in opcodes:
            if tag == 'equal':
                for line1, line2 in zip(file1_lines[i1:i2], file2_lines[j1:j2]):
                    lineno1 += 1
                    lineno2 += 1
                    ws.cell(row=row, column=1).value = lineno1
                    ws.cell(row=row, column=2).value = line1.strip('\n')
                    ws.cell(row=row, column=3).value = lineno2
                    ws.cell(row=row, column=4).value = line2.strip('\n')
                    row += 1
            elif tag == 'replace':
                self._handle_replace(ws, row, i1, i2, j1, j2, file1_lines, file2_lines, lineno1, lineno2, fill_diff, font_diff)
                row += 1
            elif tag == 'delete':
                for line1 in file1_lines[i1:i2]:
                    lineno1 += 1
                    ws.cell(row=row, column=1).value = lineno1
                    ws.cell(row=row, column=2).value = line1.strip('\n')
                    ws.cell(row=row, column=3).value = ''
                    ws.cell(row=row, column=4).value = ''
                    ws.cell(row=row, column=1).fill = fill_diff
                    ws.cell(row=row, column=2).fill = fill_diff
                    ws.cell(row=row, column=2).font = font_diff
                    row += 1
            elif tag == 'insert':
                for line2 in file2_lines[j1:j2]:
                    lineno2 += 1
                    ws.cell(row=row, column=1).value = ''
                    ws.cell(row=row, column=2).value = ''
                    ws.cell(row=row, column=3).value = lineno2
                    ws.cell(row=row, column=4).value = line2.strip('\n')
                    ws.cell(row=row, column=3).fill = fill_diff
                    ws.cell(row=row, column=4).fill = fill_diff
                    ws.cell(row=row, column=4).font = font_diff
                    row += 1

        # Adjust column widths for better readability
        ws.column_dimensions['A'].width = 6   # Line number column for Device 1
        ws.column_dimensions['B'].width = 50  # Column for Device 1 content
        ws.column_dimensions['C'].width = 6   # Line number column for Device 2
        ws.column_dimensions['D'].width = 50  # Column for Device 2 content

        # Save the workbook
        output_path = f"{self.output_filename}.xlsx"
        wb.save(output_path)
        print(f"Excel diff saved as '{output_path}'.")

    def _handle_replace(self, ws, row, i1, i2, j1, j2, file1_lines, file2_lines, lineno1, lineno2, fill_diff, font_diff):
        lines1 = file1_lines[i1:i2]
        lines2 = file2_lines[j1:j2]
        maxlen = max(len(lines1), len(lines2))
        for idx in range(maxlen):
            if idx < len(lines1):
                lineno1 += 1
                line1 = lines1[idx].strip('\n')
                line1_num = lineno1
            else:
                line1 = ''
                line1_num = ''
            if idx < len(lines2):
                lineno2 += 1
                line2 = lines2[idx].strip('\n')
                line2_num = lineno2
            else:
                line2 = ''
                line2_num = ''
            ws.cell(row=row, column=1).value = line1_num
            ws.cell(row=row, column=2).value = line1
            ws.cell(row=row, column=3).value = line2_num
            ws.cell(row=row, column=4).value = line2
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = fill_diff
            if line1:
                ws.cell(row=row, column=2).font = font_diff
            if line2:
                ws.cell(row=row, column=4).font = font_diff

    def generate_html_diff(self, file1_lines, file2_lines):
        # Generate the initial HTML diff
        differ = difflib.HtmlDiff(wrapcolumn=80)
        html_diff = differ.make_file(
            file1_lines, file2_lines, fromdesc='File 1', todesc='File 2'
        )

        # Customize the CSS to change the text color of highlighted lines
        custom_css = f'''
        <style type="text/css">
            table.diff {{font-family:Courier; border:medium;}}
            .diff_header {{background-color:#e0e0e0}}
            td.diff_header {{text-align:right}}
            .diff_next {{background-color:#f0f0f0}}
            .diff_add {{
                background-color:{self.highlight_color_html};
                font-weight:bold;
                color:{self.text_color_html};
            }}
            .diff_chg {{
                background-color:{self.highlight_color_html};
                font-weight:bold;
                color:{self.text_color_html};
            }}
            .diff_sub {{
                background-color:{self.highlight_color_html};
                font-weight:bold;
                color:{self.text_color_html};
            }}
        </style>
        '''

        # Replace the existing <style>...</style> section with the custom CSS
        html_diff = re.sub(
            r'<style type="text/css">.*?</style>',
            custom_css,
            html_diff,
            flags=re.DOTALL
        )

        # Save the modified HTML diff to a file
        output_path = f"{self.output_filename}.html"
        with open(output_path, 'w') as output_file:
            output_file.write(html_diff)

        print(f"HTML diff saved as '{output_path}'.")

    def compare_files(self, output_format='both'):
        # Read the files
        file1_lines = self.read_file(self.file1)
        file2_lines = self.read_file(self.file2)

        # Output formats: excel, html, or both
        if output_format == 'excel' or output_format == 'both':
            self.generate_excel_diff(file1_lines, file2_lines)

        if output_format == 'html' or output_format == 'both':
            self.generate_html_diff(file1_lines, file2_lines)