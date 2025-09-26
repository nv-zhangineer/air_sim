"""
Change Log

10/1/17 initial release
1/1/17 combined sheet and sheet_data into dictionary
10/17/19 added a line to skip the row if column-A is empty for that row

"""

import logging
import openpyxl
import openpyxl.utils
import sys


class ReadExcel(object):
    __slots__ = ['excel_path', 'sheet_names', 'headers']

    def __init__(self, excel_path):
        """
        Read the Excel sheet and obtain data for the sheet user requested.
            If no sheet name is specified, it will default to "sheet1" which is always the case for a new workbook
        :param excel_path: file name for the Excel file
        """
        self.excel_path = excel_path
        self.sheet_names = self._get_sheet_names()

    def _get_sheet_names(self):
        """Get list of Excel Sheet names, excluding hidden ones"""
        sheet_names = []
        wb = openpyxl.load_workbook(self.excel_path)

        for sheet in wb.sheetnames:
            logging.debug(f"Processing Sheet : {sheet}")
            if wb[sheet].sheet_state == "visible":
                logging.debug(f"Adding visible sheet to list: {sheet}")
                sheet_names.append(sheet)

        return sheet_names

    def get_excel_column_headers(self, sheet_name):
        """Read first row of an Excel Sheet to get the column headers"""
        wb = openpyxl.load_workbook(self.excel_path, data_only=True, read_only=True)
        sheet = wb.get_sheet_by_name(sheet_name)
        try:
            headers = []
            for column in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=1, column=column).value
                if cell_value:
                    headers.append(cell_value.strip())
            return headers
        except AttributeError as e:
            print(f"Detected errors in sheet {sheet_name}, potentially a formatting issue", e)

    def excel_generate_line(self, sheet_name, start_row=2):
        """
        This is the generator version of the function above, this was built to improve document loading time
        since we have no desire to keep these data but only to load them and push the configuration.
        """
        wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        ws = wb[sheet_name]
        logging.info("processing sheet {}....".format(sheet_name))
        for row in range(start_row, ws.max_row + 1):
            try:
                row_c1 = ws["A" + str(row)]
                if row_c1.value is None:
                    break
                elif row_c1:  # If A[ROW] has a value, row is a number between 2 and max row number in the sheet
                    line = dict()
                    headers = self.get_excel_column_headers(sheet_name)
                    for header in headers:
                        # Use openpyxl function to get column letter based on
                        # Header position in the Headers List plus 1
                        # Use Column Letter + Row Number to get cell value
                        col = openpyxl.utils.get_column_letter(headers.index(header) + 1)

                        ''' Add support for detecting hidden column once this is fixed
                        if not ws.column_dimensions[col].hidden: # This line does not work due to bug
                        Bug tracked here https://foss.heptapod.net/openpyxl/openpyxl/-/issues/1711
                        '''

                        cell_value = ws[col + str(row)].value
                        if isinstance(cell_value, str):  # for python3, use : str instead of basestring
                            # Convert value text from Unicode to ASCII
                            cell_value = cell_value.encode('utf-8').decode('ascii', 'ignore')
                            cell_value = cell_value.strip()  # remove white space
                        elif isinstance(cell_value, int):
                            cell_value = cell_value
                        elif cell_value is None:
                            cell_value = ''
                        else:
                            logging.info("no value found in cell: ", ws[col + str(row)].coordinate)
                        line[header] = cell_value
                    yield line
            except IndexError:
                break
