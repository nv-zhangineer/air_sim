import unittest
from openpyxl import load_workbook
import os
import sys

current_dir = os.path.dirname(os.path.realpath(__file__))
parent_dir = os.path.dirname(current_dir)
sys.path.append(parent_dir)


from util.diff_file import FileComparer

# Assuming generate_excel_diff is already defined in the imported module
# If not, import it like so: from your_module import generate_excel_diff

class TestGenerateExcelDiff(unittest.TestCase):

    def setUp(self):
        # Sample lines for file1 and file2, simulating configurations
        self.file_compare = FileComparer("tests/test_data/diff_file1.txt",
                                    "tests/test_data/diff_file2.txt",
                                    "tests/test_data/test_diff_output")
        with open("tests/test_data/diff_file1.txt", "r") as file1:
            self.file1_lines = file1.readlines()
        with open("tests/test_data/diff_file2.txt", "r") as file1:
            self.file2_lines = file1.readlines()
        # The output Excel file to be generated
        self.output_file = "tests/test_data/test_diff_output"

    @staticmethod
    def remove_empty_lines(lines):
        """Helper function to remove empty lines from the list of lines."""
        return [line for line in lines if line.strip() != ""]

    @staticmethod
    def remove_empty_rows(sheet, column):
        """Helper function to remove empty rows from a specified column in the Excel sheet."""
        non_empty_rows = []
        for row in sheet.iter_rows(min_row=2, min_col=column, max_col=column):
            if row[0].value is not None and str(row[0].value).strip() != "":
                non_empty_rows.append(row[0].value)
        return non_empty_rows

    def test_file1_lines_match(self):
        # Run the generate_excel_diff function
        self.file_compare.generate_excel_diff(self.file1_lines, self.file2_lines)

        # Load the Excel file using openpyxl
        workbook = load_workbook(f"{self.output_file}.xlsx")
        sheet = workbook.active

        # Remove empty lines from file1
        non_empty_file1_lines = self.remove_empty_lines(self.file1_lines)

        # Remove empty rows from column B (Device 1)
        non_empty_excel_file1_lines = self.remove_empty_rows(sheet, 2)
        # Assert that the number of non-empty lines in file1 matches the number of non-empty rows in column B
        self.assertEqual(len(non_empty_file1_lines), len(non_empty_excel_file1_lines),
                         "The number of non-empty lines from file1 should match the number of non-empty rows in column B.")

    def test_file2_lines_match(self):
        # Run the generate_excel_diff function
        self.file_compare.generate_excel_diff(self.file1_lines, self.file2_lines)

        # Load the Excel file using openpyxl
        workbook = load_workbook(f"{self.output_file}.xlsx")
        sheet = workbook.active

        # Remove empty lines from file2
        non_empty_file2_lines = self.remove_empty_lines(self.file2_lines)

        # Remove empty rows from column D (Device 2)
        non_empty_excel_file2_lines = self.remove_empty_rows(sheet, 4)
        # Assert that the number of non-empty lines in file2 matches the number of non-empty rows in column D
        self.assertEqual(len(non_empty_file2_lines), len(non_empty_excel_file2_lines),
                         "The number of non-empty lines from file2 should match the number of non-empty rows in column D.")

    def tearDown(self):
        # Clean up by removing the Excel file after the test runs
        if os.path.exists(f"{self.output_file}.xlsx"):
            os.remove(f"{self.output_file}.xlsx")
